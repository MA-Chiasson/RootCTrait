"""Multi-batch pipeline: processes one or several batches in a single run.

For each requested batch: load each mask -> skeleton -> decomposition ->
DECONTAMINATION -> trait EXTRACTION -> one Excel table per batch.

Expected folder layout (relative to the script folder):
    data/<batch>/      the masks of this batch
    results/<batch>/   output: traits_<batch>.xlsx, figures/, checkpoint

Config: params.txt (same folder). The processing settings (voxel, pruning,
decontamination) are SHARED by all batches, which keeps results comparable for
GWAS. Each batch only has its file-name pattern and, if needed, a specific
axis_order (empty by default: orientation is handled by the format).

Batch definition in params.txt:
    BATCH <name> | <file_pattern> | <axis_order>
e.g.:
    BATCH batch1 | sample_{name}.mat    |
    BATCH batch2 | scan{name}.tif        |
The 3rd field (axis_order) is optional; empty = no manual permutation.

Batches to process: BATCHES=ALL (all defined) or BATCHES=batch1,batch2
"""
import warnings; warnings.filterwarnings('ignore')
import os, re, time, json, math, glob
import multiprocessing as mp
import numpy as np
import scipy.ndimage as ndimage
from skimage.morphology import skeletonize
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from rootctrait.graph_extraction import prune_skeleton
from rootctrait.root_decomposition import decompose_root_system
from rootctrait.decontamination import decontaminate, keep_base_component
from rootctrait.root_traits_full import compute_all_traits
from rootctrait.detection_hypocotyle import collar_and_hypocotyl
from rootctrait.io_volume import load_volume

# ============================ CONFIGURATION ============================

def load_params(path):
    """Read params.txt. Returns (dict of simple keys, list of batches).
    A batch is a line 'BATCH name | pattern | axis_order'."""
    p = {}
    batches = []
    if os.path.exists(path):
        with open(path, encoding='utf-8-sig') as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith(';'):
                    continue
                if line.upper().startswith('BATCH '):
                    # BATCH name | pattern | axis_order(optional)
                    body = line[6:]
                    parts = [x.strip() for x in body.split('|')]
                    bname = parts[0]
                    pattern = parts[1] if len(parts) > 1 else ''
                    ao = parts[2] if len(parts) > 2 else ''
                    axis_order = tuple(int(x) for x in ao.split(',')) if ao else None
                    batches.append({'name': bname, 'pattern': pattern, 'axis_order': axis_order})
                elif '=' in line:
                    k, v = line.split('=', 1)
                    p[k.strip().upper()] = v.strip()
    return p, batches


PARAMS, BATCH_DEFS = load_params(os.environ.get('PARAMS', 'params.txt'))

DATA_ROOT = PARAMS.get('DATA_ROOT', 'data')
RESULTS_ROOT = PARAMS.get('RESULTS_ROOT', 'results')
VOXEL_SIZE = tuple(float(x) for x in PARAMS.get('VOXEL', '0.39,0.39,0.2').split(','))
PRUNE_VOX = int(PARAMS.get('PRUNE_VOX', '5'))
MIN_SEG_LEN_MM = float(PARAMS.get('MIN_SEG_LEN_MM', '2.0'))
BC_MIN = int(PARAMS.get('BC_MIN', '3'))
LIN_MAX = float(PARAMS.get('LIN_MAX', '0.7'))
LEN_MAX = float(PARAMS.get('LEN_MAX', '15'))
DROP_ORPHANS = PARAMS.get('DROP_ORPHANS', '1').lower() not in ('0', 'false', 'no')
SAVE_FIGURES = PARAMS.get('SAVE_FIGURES', '1').lower() not in ('0', 'false', 'no')
TIMEOUT = int(PARAMS.get('TIMEOUT', '1800'))

# Which batches to process
_batch = PARAMS.get('BATCHES', 'ALL').strip()
if _batch.upper() == 'ALL':
    ACTIVE_BATCHES = BATCH_DEFS
else:
    requested = [b.strip() for b in _batch.split(',') if b.strip()]
    ACTIVE_BATCHES = [b for b in BATCH_DEFS if b['name'] in requested]
# ======================================================================

COLS = [('LRP', 'cm', .1), ('TRL', 'cm', .1), ('LTRL', 'cm', .1), ('MLRL', 'cm', .1),
        ('NRL', 'compte', 1), ('NRL_court_<5', 'compte', 1), ('NRL_moyen_5_15', 'compte', 1),
        ('NRL_long_>15', 'compte', 1), ('PM', 'cm', .1), ('D50', 'cm', .1), ('D95', 'cm', .1),
        ('WX', 'cm', .1), ('WZ', 'cm', .1), ('LM', 'cm', .1), ('W25', 'cm', .1), ('W50', 'cm', .1),
        ('W75', 'cm', .1), ('RLP', 'ratio', 1), ('ANGsys', 'deg', 1), ('ACRL', 'deg', 1),
        ('ANGO2', 'deg', 1), ('ANGO2_sd', 'deg', 1), ('ANGO2_init', 'deg', 1),
        ('CHV', 'cm3', .001), ('VRT', 'cm3', .001), ('SRT', 'cm2', .01), ('IC', 'ratio', 1),
        ('SRL', 'cm/cm3', 100), ('NT', 'compte', 1), ('NBP', 'compte', 1), ('MaxO', 'compte', 1),
        ('DR', 'nb/cm', 1), ('NTR', 'compte', 1), ('IBD', 'cm', .1), ('DRP', 'mm', 1),
        ('DRS', 'mm', 1), ('DMAX', 'mm', 1), ('DD_cv', 'ratio', 1), ('TAPER', 'frac/cm', 1),
        ('TOR', 'ratio', 1)]


def detect_base(sk, edt, voxel_size):
    """Collar detection: THICKEST point in the upper layer of the skeleton (the
    shallowest 20% in depth). The true collar is the widest structure at the top of
    the system; anchoring there straightens and re-centers the pivot. This is more
    robust than a multi-criteria (depth/thickness/centrality) rule, especially when
    the primary root is thin, a frequent case that used to anchor the collar on an
    off-center thin root."""
    sc = np.argwhere(sk)
    if len(sc) == 0:
        raise ValueError("Empty skeleton")
    seuil = np.percentile(sc[:, 0], 20)
    haut = sc[sc[:, 0] <= seuil]
    rad = edt[haut[:, 0], haut[:, 1], haut[:, 2]]
    return haut[np.argmax(rad)]


def make_figure(name, kept, removed_segs, primary_path, fig_dir, hypocotyl=None,
                orphans=None, base=None, base2=None):
    try:
        import plotly.graph_objects as go
    except ImportError:
        return False
    vs = np.asarray(VOXEL_SIZE)

    def lines(segs):
        x, y, z = [], [], []
        for s in segs:
            C = s['coords'] * vs
            x += list(C[:, 1]) + [None]; y += list(C[:, 2]) + [None]; z += list(C[:, 0]) + [None]
        return x, y, z

    fig = go.Figure()
    xk, yk, zk = lines(kept)
    fig.add_trace(go.Scatter3d(x=xk, y=yk, z=zk, mode='lines',
                               line=dict(color='#1f77b4', width=2), name='kept laterals'))
    xr, yr, zr = lines(removed_segs)
    fig.add_trace(go.Scatter3d(x=xr, y=yr, z=zr, mode='lines',
                               line=dict(color='#d62728', width=2), name='removed (pollution)'))
    if hypocotyl:
        xh, yh, zh = lines(hypocotyl)
        fig.add_trace(go.Scatter3d(x=xh, y=yh, z=zh, mode='lines',
                                   line=dict(color='#ff7f0e', width=3), name='hypocotyl (excluded)'))
    if orphans:
        xo, yo, zo = lines(orphans)
        fig.add_trace(go.Scatter3d(x=xo, y=yo, z=zo, mode='lines',
                                   line=dict(color='#bbbbbb', width=2), name='orphans (removed)'))
    if primary_path is not None and len(primary_path):
        P = np.asarray(primary_path) * vs
        fig.add_trace(go.Scatter3d(x=P[:, 1], y=P[:, 2], z=P[:, 0], mode='lines',
                                   line=dict(color='black', width=6), name='pivot'))
    if base is not None:
        b = np.asarray(base) * vs
        fig.add_trace(go.Scatter3d(x=[b[1]], y=[b[2]], z=[b[0]], mode='markers',
                                   marker=dict(color='green', size=6), name='collar (thick point)'))
    if base2 is not None:
        b2 = np.asarray(base2) * vs
        fig.add_trace(go.Scatter3d(x=[b2[1]], y=[b2[2]], z=[b2[0]], mode='markers',
                                   marker=dict(color='#7b2fbe', size=7, symbol='diamond'),
                                   name='raised collar (RP start)'))
    fig.update_layout(scene=dict(
        xaxis_title='X (mm)', yaxis_title='Z (mm)', zaxis_title='Y depth (mm)',
        zaxis=dict(autorange='reversed')), title=name)
    os.makedirs(fig_dir, exist_ok=True)
    fig.write_html(os.path.join(fig_dir, f"{name}.html"), include_plotlyjs='cdn')
    return True


# The current batch is passed to the workers via module variables
# The current batch context (data_dir, pattern, axis_order, fig_dir) is passed
# explicitement aux workers, pour compatibilite avec le mode 'spawn' de Windows.



def _extend_pivot(prim, sk, base, base2, voxel_size, marge=15):
    """Extend the pivot upward to the raised collar base2 following the REAL skeleton
    path: shortest path (Dijkstra) between base2 and base, computed on a small local
    window around both collars. The primary root is thus drawn in black continuously
    from the old to the new collar, following the root, without a loop. If no path
    exists (collars not connected in the window), a straight segment is used."""
    from scipy.spatial import cKDTree
    from scipy.sparse import coo_matrix
    from scipy.sparse.csgraph import dijkstra
    prim = np.asarray(prim)
    base = np.asarray(base)
    base2 = np.asarray(base2)
    vs = np.asarray(voxel_size)

    def _droit():
        if len(prim) == 0:
            return base2.reshape(1, 3)
        if np.array_equal(base2, prim[0]):
            return prim
        return np.vstack([base2.reshape(1, 3), prim])

    if np.array_equal(base, base2):
        return prim if len(prim) else base2.reshape(1, 3)
    lo = np.maximum(np.minimum(base, base2) - marge, 0)
    hi = np.maximum(base, base2) + marge
    sl = tuple(slice(int(lo[i]), int(hi[i]) + 1) for i in range(3))
    sub = np.zeros_like(sk)
    sub[sl] = sk[sl]
    vox = np.argwhere(sub)
    if len(vox) < 2:
        return _droit()
    tree = cKDTree(vox)
    pairs = tree.query_pairs(r=np.sqrt(3) + 1e-6, output_type='ndarray')
    if len(pairs) == 0:
        return _droit()
    d = np.linalg.norm((vox[pairs[:, 0]] - vox[pairs[:, 1]]) * vs, axis=1)
    n = len(vox)
    rows = np.concatenate([pairs[:, 0], pairs[:, 1]])
    cols = np.concatenate([pairs[:, 1], pairs[:, 0]])
    G = coo_matrix((np.concatenate([d, d]), (rows, cols)), shape=(n, n)).tocsr()

    def _idx(pt):
        m = np.where((vox == pt).all(1))[0]
        return int(m[0]) if len(m) else int(np.argmin(((vox - pt) ** 2).sum(1)))
    i2, i1 = _idx(base2), _idx(base)
    dist, pred = dijkstra(G, indices=i2, return_predecessors=True)
    if not np.isfinite(dist[i1]):
        return _droit()
    path = [i1]; k = i1
    while k != i2 and pred[k] >= 0:
        k = int(pred[k]); path.append(k)
    route = vox[path[::-1]]  # base2 -> base
    if len(prim) == 0:
        return route
    return np.vstack([route, prim])


def process(name, ctx):
    path = os.path.join(ctx['data_dir'], ctx['pattern'].format(name=name))
    V = load_volume(path, axis_order=ctx['axis_order'])
    BW = V > (0.5 * np.max(V))
    c = np.argwhere(BW); mn = c.min(0); mx = c.max(0); m = 6
    BW = BW[tuple(slice(max(0, mn[i] - m), mx[i] + m) for i in range(3))]
    edt = ndimage.distance_transform_edt(BW, sampling=VOXEL_SIZE)
    sk = skeletonize(BW).astype(bool)
    sk = prune_skeleton(sk, min_branch_length_vox=PRUNE_VOX)
    base = detect_base(sk, edt, VOXEL_SIZE)
    segs, prim, voxels, border, bnode = decompose_root_system(
        sk, base, list(VOXEL_SIZE), dist_map=edt,
        min_seg_len_mm=MIN_SEG_LEN_MM, crown_exclude_mm=0.0)
    n_brut = len(segs)
    kept, removed_segs, feats = decontaminate(segs, voxels, VOXEL_SIZE, base=base,
                                              bc_min=BC_MIN, lin_max=LIN_MAX, len_max=LEN_MAX,
                                              drop_orphans=DROP_ORPHANS)
    n_ret = len(removed_segs)
    # --- Bounded collar + hypocotyl (refined version) ---
    # base2 = collar raised along the thick column up to the hypocotyl (start of the
    # pivot at the true top of the fleshy base). hypo_ids = vertical column + the
    # branches hanging high on the stem; horizontal roots at collar level stay roots.
    base2, hypo_ids, gain, column_pts = collar_and_hypocotyl(sk, edt, kept, base, VOXEL_SIZE)
    roots = [s for s in kept if s['seg_id'] not in hypo_ids]
    hypocotyl = [s for s in kept if s['seg_id'] in hypo_ids]
    # Orphan cleanup: a single call to the existing function, after removing the
    # hypocotyl, to drop detached fragments. Memory guard on very large systems
    # (keep_base_component builds a tree of all voxels).
    orphans = []
    nvox_roots = sum(len(s['coords']) for s in roots)
    if roots and nvox_roots <= 90000:
        try:
            roots, orphans = keep_base_component(roots, base)
        except Exception:
            orphans = []
    # Pivot extended up to the raised collar, so that the pivot length (LRP) starts
    # at base2 (the true start of the primary root).
    prim2 = _extend_pivot(prim, sk, base, base2, VOXEL_SIZE)
    skv_clean = np.vstack([s['coords'] for s in roots]) if roots else voxels
    if SAVE_FIGURES:
        make_figure(name, roots, removed_segs, prim2, ctx['fig_dir'],
                    hypocotyl=hypocotyl, orphans=orphans, base=base, base2=base2)
    # base2 = reference collar for the traits (depths, LRP, angles).
    T = compute_all_traits(roots, prim2, base2, BW, edt, VOXEL_SIZE, skv_clean)
    return n_brut, n_ret, T


def write_excel(results, out_xlsx, title):
    wb = Workbook(); ws = wb.active; ws.title = "Traits"; F = "Arial"
    ncol = len(COLS) + 4
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    ws.cell(1, 1, title)
    ws.cell(1, 1).font = Font(name=F, bold=True, size=12, color="FFFFFF")
    ws.cell(1, 1).fill = PatternFill("solid", start_color="1F3864")
    ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22
    hdr = ["ID", "n_brut", "n_retire", "%retire"] + [c[0] for c in COLS]
    uni = ["", "", "", ""] + [c[1] for c in COLS]
    for j, h in enumerate(hdr, 1):
        cc = ws.cell(2, j, h); cc.font = Font(name=F, bold=True, color="FFFFFF")
        cc.fill = PatternFill("solid", start_color="2E5496")
        cc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for j, u in enumerate(uni, 1):
        cc = ws.cell(3, j, u); cc.font = Font(name=F, italic=True, size=8, color="D9D9D9")
        cc.fill = PatternFill("solid", start_color="2E5496")
        cc.alignment = Alignment(horizontal="center")
    thin = Side(style="thin", color="E0E0E0"); bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    r = 4
    for name, nb, nr, T in results:
        band = "F5F8FC" if r % 2 == 0 else "FFFFFF"
        vals = [name, nb, nr, (round(100 * nr / nb, 1) if nb else None)]
        for j, v in enumerate(vals, 1):
            cc = ws.cell(r, j, v); cc.border = bd; cc.fill = PatternFill("solid", start_color=band)
            if j == 1:
                cc.font = Font(name=F, bold=True)
        for j, (key, unit, fac) in enumerate(COLS, 5):
            tv = T.get(key) if T is not None else None
            cc = ws.cell(r, j); cc.border = bd; cc.fill = PatternFill("solid", start_color=band)
            if tv is None or (isinstance(tv, float) and not math.isfinite(tv)):
                cc.value = None
            else:
                v = float(tv) * fac
                cc.value = round(v, 3 if abs(v) < 100 else 1)
                cc.number_format = '0.000' if abs(v) < 100 else '0.0'
        r += 1
    ws.freeze_panes = "E4"
    for col, w in (('A', 8), ('B', 8), ('C', 9), ('D', 8)):
        ws.column_dimensions[col].width = w
    for j in range(5, ncol + 1):
        ws.column_dimensions[get_column_letter(j)].width = 10
    wb.save(out_xlsx)


def _clean_T(T):
    if T is None:
        return None
    out = {}
    for k, v in T.items():
        if v is None:
            out[k] = None; continue
        try:
            fv = float(v)
        except (TypeError, ValueError):
            out[k] = None; continue
        if not math.isfinite(fv):
            out[k] = None
        elif fv == int(fv):
            out[k] = int(fv)
        else:
            out[k] = fv
    return out


def load_store(path):
    recs = {}
    if os.path.exists(path):
        with open(path, encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    r = json.loads(line); recs[r['name']] = r
                except Exception:
                    pass
    return recs


def append_store(path, rec):
    with open(path, 'a', encoding='utf-8') as f:
        f.write(json.dumps(rec) + '\n')


def _worker(name, ctx, q):
    try:
        q.put(('ok', process(name, ctx)))
    except Exception as e:
        q.put(('err', repr(e)))


def run_with_timeout(name, ctx, timeout):
    q = mp.Queue()
    p = mp.Process(target=_worker, args=(name, ctx, q))
    p.start()
    p.join(timeout)
    if p.is_alive():
        p.terminate(); p.join()
        return ('timeout', None)
    try:
        return q.get_nowait()
    except Exception:
        return ('err', 'processus termine sans renvoyer de resultat')


def list_samples(data_dir, pattern):
    pre, suf = pattern.split('{name}')
    names = []
    for fp in glob.glob(os.path.join(data_dir, pre + '*' + suf)):
        b = os.path.basename(fp)
        names.append(b[len(pre):len(b) - len(suf)] if suf else b[len(pre):])
    return sorted(names, key=lambda s: int(re.sub(r'\D', '', s) or 0))


def process_batch(batch):
    bname = batch['name']
    data_dir = os.path.join(DATA_ROOT, bname)
    res_dir = os.path.join(RESULTS_ROOT, bname)
    os.makedirs(res_dir, exist_ok=True)
    out_xlsx = os.path.join(res_dir, f"traits_{bname}.xlsx")
    checkpoint = os.path.join(res_dir, 'checkpoint_traits.jsonl')
    fig_dir = os.path.join(res_dir, 'figures')

    ctx = {'data_dir': data_dir, 'pattern': batch['pattern'],
           'axis_order': batch['axis_order'], 'fig_dir': fig_dir}

    if not os.path.isdir(data_dir):
        print(f"[{bname}] folder not found: {data_dir} -- batch skipped.\n")
        return
    samples = list_samples(data_dir, batch['pattern'])
    if not samples:
        print(f"[{bname}] no file matching {batch['pattern']} in {data_dir} -- skipped.\n")
        return

    store = load_store(checkpoint)
    done = [s for s in samples if s in store]
    todo = [s for s in samples if s not in store]
    print(f"=== BATCH {bname} : {len(samples)} samples | {len(done)} already done | "
          f"{len(todo)} to do | pattern={batch['pattern']} | axis_order={batch['axis_order']} ===")
    for name in todo:
        t0 = time.time()
        status, payload = run_with_timeout(name, ctx, TIMEOUT)
        if status == 'ok':
            nb, nr, T = payload
            rec = {'name': name, 'n_brut': nb, 'n_ret': nr, 'T': _clean_T(T)}
            store[name] = rec
            append_store(checkpoint, rec)
            print(f"  {name:6s} raw={nb:4d} removed={nr:4d} ({100*nr/max(1,nb):3.0f}%) "
                  f"LRP={T['LRP']/10:.1f}cm NRL={T['NRL']}  [{time.time()-t0:.0f}s]", flush=True)
        elif status == 'timeout':
            print(f"  {name:6s} TIMEOUT after {TIMEOUT}s (will be retried)", flush=True)
        else:
            print(f"  {name:6s} ERREUR {payload} (sera reessaye)", flush=True)
    results = [(s, store[s]['n_brut'], store[s]['n_ret'], store[s]['T']) for s in samples if s in store]
    title = f"Root traits - {bname} - after decontamination (cleaned skeleton)"
    write_excel(results, out_xlsx, title)
    ok = len(results); fail = len(samples) - ok
    print(f"  -> {out_xlsx}  ({ok}/{len(samples)} samples)")
    if fail:
        print(f"  -> {fail} non aboutis (relancez pour les reprendre)")
    print()


def main():
    if not ACTIVE_BATCHES:
        print("No batch to process. Check BATCHES and the BATCH lines in params.txt.")
        return
    print(f"Batches to process: {', '.join(b['name'] for b in ACTIVE_BATCHES)}\n")
    for batch in ACTIVE_BATCHES:
        process_batch(batch)
    print("Done.")


if __name__ == "__main__":
    mp.freeze_support()
    main()
